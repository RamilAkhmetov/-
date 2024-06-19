import multiprocessing as mp
mp.freeze_support()
import sys
import os
import time
import re
import csv
from openpyxl import Workbook

from PyQt6 import QtCore, QtWidgets, QtGui

from AnalitMainWindow import Ui_MainWindow

import cv2
import numpy as np
from Finder import Finder_of_main_rectangle, Finder_of_circles_contours_cv2
from Checker import Checker_of_circles_contours
from Color import ColorExtracter

from Processor import CellsFinder, CellsPostprocessor
from Style import style
from PhotoProcessingWidgets import AdjustmentDialog

from cfg import Config

def normalize_path(path):
    n_path = ""
    for s in path:
        if s == '\\':
            n_path += '/'

    return n_path

def edit_path(file_path):
    # Convert the file path to the Unicode string
    file_path = file_path.encode('utf-8').decode('utf-8')

    # Check if the file exists
    if not os.path.isfile(file_path):
        print("File not found!")
        return None

    return file_path



def find_child( parent_object: QtWidgets.QWidget, child_type=None, child_name=None):
    """
    Должен быть задан либо тип, либо имя, либо и то, и другое

    :param parent_object: QtWidgets.QWidget
        виджет-родитель, у которого ищем наследника
    :param child_type:
        тип наследника
    :param child_name: str
        имя наследника
    :return:
    """
    if (child_type is not None) and (child_name is None):
        for widget in parent_object.children():
            if isinstance(widget, child_type):
                return widget
    elif (child_type is None) and (child_name is not None):
        for widget in parent_object.children():
            if widget.objectName() == child_name:
                return widget
    elif (child_type is None) and (child_name is None):
        for widget in parent_object.children():
            if isinstance(widget, child_type) and widget.objectName() == child_name:
                return widget
    else:
        return None

def from_cv2_to_pixmap(cv2_image, bgr_to_rgb=True):
    image = cv2_image.copy()
    if bgr_to_rgb:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    else:
        pass
    qt_image = QtGui.QImage(image, image.shape[1], image.shape[0], image.shape[1] * 3, QtGui.QImage.Format.Format_RGB888)
    print("ok qt_image")
    pix = QtGui.QPixmap(qt_image)
    return pix

def scale_cv2(cv2_image, width, height):
    return cv2.resize(cv2_image, (width, height), interpolation=cv2.INTER_CUBIC)


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    """
    Главное окно приложения
    Main window of application


    Аттрибуты
    Attributes
    ----------
    modes: list
        Список сцен интерфейса
        List of modes (scenes)

    current_mode: int
        Индекс текущей сцены
        Index of current mode

    Методы
    Methods
    -------
    setupMyUi()
        Определяет интерфейс всего приложения
        Sets user interface of whole application

    """
    def __init__(self):
        super(MainWindow, self).__init__()
        self.modes = ["Новое исследование", "Обработка фото",]
        self.current_mode = None
        self.setupUi(self)
        self.setWindowTitle("Chrometrica")
        self.setWindowState(QtCore.Qt.WindowState.WindowMaximized)

        self.photo_process = None
        self.setupMyUi()
        self.show()

    def setupMyUi(self):
        """
        Определяет интерфейс всего приложения
        Sets user interface of whole application
        """
        # определить меню
        # set menu
        self.setupMenu()
        self.setupui()
        #

    def setupMenu(self):
        """
        Определяет меню приложения
        Creates app menu
        """

        # создать кнопку "Новое исследование"
        # make a "Create study" button
        #create_study_button_action = QtGui.QAction("&Новое исследование", self)
        # связать с обработчиком
        # bid with slot
        #create_study_button_action.triggered.connect(self.onCreateStudyButtonClicked)

        # создать кнопку "Открыть фото"
        # make a "Open photo" button
        open_photo_button_action = QtGui.QAction("&Открыть фото", self)
        # связать с обработчиком
        # bid with slot
        open_photo_button_action.triggered.connect(self.onOpenPhotoButtonClicked)

        open_dir_button_action = QtGui.QAction("&Открыть папку", self)
        open_dir_button_action.triggered.connect(self.onOpenDirButtonClicked)

        create_samples_schema = QtGui.QAction("&Создать схему образцов", self)
        create_samples_schema.triggered.connect(self.onCreateSamplesSchemaButtonClicked)

        # создать меню у главного окна
        # create a menu of main window
        menu = self.menuBar()


        # добавить в меню раздел "Файл"
        # add "File" section to menu
        file_menu = menu.addMenu("&Файл")
        # добавить в раздел "Файл" кнопку "Новое иссследование"
        # add "Create study" button to "File" section
        #file_menu.addAction(create_study_button_action)
        # добавить в раздел "Файл" разделитель между кнопками
        # add separator to "File" section
        #file_menu.addSeparator()
        # добавить в раздел "Файл" кнопку "Открыть фото"
        # add "Open photo" to "File" section
        file_menu.addAction(open_photo_button_action)
        file_menu.addAction(open_dir_button_action)
        file_menu.addSeparator()
        file_menu.addAction(create_samples_schema)

    def setupui(self):
        """
        Задать интерфейс (все виджеты, не все показываются)
        Setting up photo processing interface on header

        """
        header_layout = QtWidgets.QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 0)
        self.header.setLayout(header_layout)


        # создать пустой виджет, который показыввется, когда ни один из режимов не включен
        # create a widget for default state
        header_empty_frame = QtWidgets.QFrame(objectName="header_empty_frame")

        # создать виджет сетки для кнопок и прочего для режима "Обработка фото"
        # make a layout for buttons and other of "Photo processing" mode
        header_photo_processing_box = QtWidgets.QFrame(objectName="header_photo_processing_box")
        # найти виджет разметки, принадлежащий блок-контейнеру для виджетов
        header_photo_processing_layout = QtWidgets.QGridLayout()
        header_photo_processing_box.setLayout(header_photo_processing_layout)

        header_photo_processing_files_combobox = QtWidgets.QComboBox(
            objectName="header_photo_processing_files_combobox")


        header_photo_processing_rectangle_button = QtWidgets.QPushButton(text="Область распознавания",
                                                                         objectName="header_photo_processing_rectangle_button")
        header_photo_processing_rectangle_button.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
                                                               QtWidgets.QSizePolicy.Policy.Maximum)
        header_photo_processing_rectangle_button.setCheckable(True)


        header_photo_processing_circles_button = QtWidgets.QPushButton(text="Контуры ячеек",
                                                                       objectName="header_photo_processing_circles_button")
        header_photo_processing_circles_button.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
                                                             QtWidgets.QSizePolicy.Policy.Maximum)
        header_photo_processing_circles_button.setCheckable(True)


        header_photo_processing_colors_button = QtWidgets.QPushButton(text="Цвета ячеек", objectName="header_photo_processing_colors_button")
        header_photo_processing_colors_button.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
                                                               QtWidgets.QSizePolicy.Policy.Maximum)
        header_photo_processing_colors_button.setCheckable(True)
        header_photo_processing_colors_button.setChecked(True)

        stages_button_group = QtWidgets.QButtonGroup(parent=header_photo_processing_box)
        stages_button_group.addButton(header_photo_processing_rectangle_button, 0)
        stages_button_group.addButton(header_photo_processing_circles_button, 1)
        stages_button_group.addButton(header_photo_processing_colors_button, 2)


        header_photo_processing_save_button = QtWidgets.QPushButton(text="Сохранить",
                                                                    objectName="header_photo_processing_save_button")
        header_photo_processing_save_button.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
                                                            QtWidgets.QSizePolicy.Policy.Maximum)

        header_photo_processing_layout.addWidget(header_photo_processing_files_combobox, 0, 0)
        header_photo_processing_layout.setAlignment(header_photo_processing_files_combobox,
                                                    QtCore.Qt.AlignmentFlag.AlignLeft)
        header_photo_processing_layout.addWidget(header_photo_processing_rectangle_button, 0, 1)
        header_photo_processing_layout.addWidget(header_photo_processing_circles_button, 0, 2)
        header_photo_processing_layout.addWidget(header_photo_processing_colors_button, 0, 3)
        header_photo_processing_layout.addWidget(header_photo_processing_save_button, 0, 4)



        # режим папки
        header_dir_processing_box = QtWidgets.QFrame(objectName="header_dir_processing_box")
        header_dir_processing_layout = QtWidgets.QGridLayout()
        header_dir_processing_box.setLayout(header_dir_processing_layout)

        header_dir_processing_files_combobox = QtWidgets.QComboBox(
            objectName="header_dir_processing_files_combobox")


        # создать схему образцов
        header_schema_making_box = QtWidgets.QFrame(objectName="header_schema_making_box")
        header_schema_making_layout = QtWidgets.QGridLayout()
        header_schema_making_box.setLayout(header_schema_making_layout)








        # создать виджет для размещения блоков виджетов, отвечающих определенным режимам
        # make a layout for stacking widgets

        header_stacked_widget = QtWidgets.QStackedWidget()

        # добавить пустой блок в стек-виджет первым
        header_stacked_widget.addWidget(header_empty_frame)
        header_stacked_widget.addWidget(header_photo_processing_box)


        #stacked_widget.setSizePolicy(QtWidgets.QSizePolicy.Policy.MinimumExpanding, QtWidgets.QSizePolicy.Policy.MinimumExpanding)

        header_layout.addWidget(header_stacked_widget)



        print("ok")


        right_sidebar_layout = QtWidgets.QVBoxLayout()
        right_sidebar_layout.setContentsMargins(0, 0, 0, 0)
        self.right_sidebar.setLayout(right_sidebar_layout)
        right_sidebar_stacked_widget = QtWidgets.QStackedWidget()
        right_sidebar_empty_frame = QtWidgets.QFrame(objectName="right_sidebar_empty_frame")
        right_sidebar_photo_processing_stacked_box = QtWidgets.QStackedWidget(objectName="right_sidebar_photo_processing_box")
        right_sidebar_photoprocessing_info_frame = QtWidgets.QFrame(
            objectName="right_sidebar_photoprocessing_info_frame")
        right_sidebar_photoprocessing_contours_frame = QtWidgets.QFrame(
            objectName="right_sidebar_photoprocessing_contours_frame")
        right_sidebar_photoprocessing_contours_layout = QtWidgets.QGridLayout(objectName="right_sidebar_photoprocessing_contours_layout")
        right_sidebar_photoprocessing_contours_edit_button_icon = QtGui.QIcon("icons/icons8-select-cursor-24.png")
        right_sidebar_photoprocessing_contours_edit_button = QtWidgets.QPushButton(
            right_sidebar_photoprocessing_contours_edit_button_icon, "Редактировать", objectName="right_sidebar_photoprocessing_contours_edit_button")
        right_sidebar_photoprocessing_contours_apply_button = QtWidgets.QPushButton("Применить", objectName="right_sidebar_photoprocessing_contours_apply_button")
        right_sidebar_photoprocessing_contours_radius_fraction_input = QtWidgets.QDoubleSpinBox(objectName="right_sidebar_photoprocessing_contours_radius_fraction_input")
        right_sidebar_photoprocessing_contours_layout.addWidget(right_sidebar_photoprocessing_contours_edit_button, 1,
                                                                1)
        right_sidebar_photoprocessing_contours_layout.addWidget(right_sidebar_photoprocessing_contours_apply_button, 1,
                                                                2)
        right_sidebar_photoprocessing_contours_layout.addWidget(
            right_sidebar_photoprocessing_contours_radius_fraction_input, 2, 1)
        right_sidebar_photoprocessing_contours_frame.setLayout(right_sidebar_photoprocessing_contours_layout)
        right_sidebar_photoprocessing_colors_frame = QtWidgets.QFrame(
            )
        right_sidebar_photo_processing_stacked_box.addWidget(right_sidebar_photoprocessing_info_frame)
        right_sidebar_photo_processing_stacked_box.addWidget(right_sidebar_photoprocessing_contours_frame)
        right_sidebar_photo_processing_stacked_box.addWidget(right_sidebar_photoprocessing_colors_frame)
        right_sidebar_photo_processing_stacked_box.setCurrentIndex(2)
        right_sidebar_stacked_widget.addWidget(right_sidebar_empty_frame)
        right_sidebar_stacked_widget.addWidget(right_sidebar_photo_processing_stacked_box)
        right_sidebar_stacked_widget.setHidden(True)
        right_sidebar_layout.addWidget(right_sidebar_stacked_widget)
        # right_sidebar_photo_processing_layout = QtWidgets.QGridLayout()
        # right_sidebar_photo_processing_box.setLayout(right_sidebar_photo_processing_layout)
        # right_sidebar_photo_processing_save_table_button = QtWidgets.QPushButton(text="Сохранить таблицу",
        #                                                             objectName="right_sidebar_photo_processing_save_table_button")
        # right_sidebar_photo_processing_save_table_button.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
        #                                                   QtWidgets.QSizePolicy.Policy.Maximum)
        # right_sidebar_photo_processing_save_image_button = QtWidgets.QPushButton(text="Сохранить картинки",
        #                                                                          objectName="right_sidebar_photo_processing_save_image_button")
        # right_sidebar_photo_processing_save_image_button.setSizePolicy(QtWidgets.QSizePolicy.Policy.Maximum,
        #                                                                QtWidgets.QSizePolicy.Policy.Maximum)
        print("Main Window setup UI ")





    def onCreateSamplesSchemaButtonClicked(self):
        pass

    def onOpenDirButtonClicked(self):
        pass

    def onOpenPhotoButtonClicked(self):
        """
        Слот для кнопки "Открыть фото"
        "Open photo" button slot
        """

        if self.photo_process is not None:
            self.photo_process.close()
            self.photo_process = None
            print("closed")
        self.photo_process = PhotoProcessing(self, self.header, self.left_sidebar,
                                             self.right_sidebar, self.frame)




    def updateHeader(self):
        pass

class OpenPhotoDialog(QtWidgets.QFileDialog):
    last_directory = ""
    def __init__(self):
        super(OpenPhotoDialog, self).__init__()

class SamplesSchemaMaker:
    def __init__(self, ):
        pass

class DirProcessing(QtCore.QObject):
    def __init__(self, filenames_list, header, left_sidebar, right_sidebar, frame):
        super(DirProcessing, self).__init__()
        self.header = header

        self.left_sidebar = left_sidebar
        self.right_sidebar = right_sidebar
        self.frame = frame

        self.filenames = filenames_list
        self.photos_count = len(self.filenames)



class PhotoProcessing(QtCore.QObject):

    """
    Обработчик фото
    Photo processor

    Получает путь к файлу, выполняет поиск ячеек и вычленяет их цвета.
    Дает возможность сохранить таблицу с rgb кодами для ячеек.

    Attributes
    ----------
    header: QtWidgets.QWidget
        Ссылка на верхний блок окна, в котором будут располагаться кнопки и прочее
        Header object

    left_sidebar: QtWidgets.QWidget
        Ссылка на правый блок окна, в котором будут располагаться кнопки и прочее
        Header object

    right_sidebar: QtWidgets.QWidget
        Ссылка на левый блок окна, в котором будут располагаться кнопки и прочее
        Header object

    filenames: [str, ]
        Список файлов
        File paths of image

    initial_image: cv2.Image
        Исходное изображение
        Initial image got from given path

    work_image: cv2.Image
        Рабочая область изображения, должна определяться автоматически
        Work area of image

    contours: np.array
        Контуры ячеек
        Contours of cells

    circles_centers: [(int, int)]
        Круги, описывающие полученные контуры ячеек
        Approximating circles of cells

    circles_radius: int
        Радиус описывающих окружностей (ячеек на изображении)
        Radius of approximating circles

    circles_objects: [QtWidgets.QGraphicsItem]
        Виджеты кругов, которыми можно будет манипулировать
        List of circles-widgets



    """
    def __init__(self, main_window, header, left_sidebar, right_sidebar, frame):
        """

        :param filenames_list: [str, ]
            Список выбранных файлов
            File paths

        :param header: QtWidgets.QFrame
            Ссылка на верхний блок, через которую можно бует менять интерфейс

        :param left_sidebar: QtWidgets.QFrame

        :param right_sidebar: QtWidgets.QFrame

        :param frame: QtWidgets.QFrame
        """
        super(PhotoProcessing, self).__init__()
        # виджеты, которыми будем манипулировать извне. Будем размещать кнопки и прочее в эти контейнеры
        self.config = Config()
        self.main_window = main_window
        self.header = header
        self.left_sidebar = left_sidebar
        self.right_sidebar = right_sidebar
        self.frame = frame

        # переменные интерфейса
        self.current_image_index = 0  # int: текущее фото
        self.current_stage_indices = 2  # [int]: список чисел текущих стадий




        # виджеты
        #self.stages_button_group = None  # QtWidgets.QButtonGroup: "Область распознавания", "Контуры", "Цвета"
        #self.frame_stacked_widget = None  # QtWidgets.QStackedWidget: стек-виджет для фото
        #self.stage_stacked_widgets = None  # [QtWidgets.QStackedWidget]: список стек-виджетов для стадий обработки





        self.photos_selected = False
        self.process_adjusted = False
        # аттрибуты для обработки фото
        self.filenames = [] # [str]
        self.photos_count = None

        self.modes = dict.fromkeys(self.config.modes)
        for m in self.modes.keys():
            self.modes[m] = []

        self.conf = None
        self.max_det = None
        self.iou = None
        self.max_rows = None
        self.max_columns = None

        self.images_modes = {}
        self.wrong_modes = []
        self.wrong_images = []
        self.results = {}
        self.cellspostprocess_data = {}

        # contours
        self.right_sidebar_photoprocessing_contours_radius_fraction_input = None
        self.contours_pixmap_item = None
        self.radius_fractions = []

        self.initial_images = {}  # []
        self.main_rectangle_boxes = []  # []
        self.main_rectangle_images = []  # []
        self.contours = []  # [[]]
        self.circles_centers = []  # [[]]
        self.circles_radius = []  # []
        self.circles_objects = []  # [[]]
        self.colors = []

        #self.processImage()
        self.selectFiles()
        #self.preprocessImages()


    def selectFiles(self):
        open_photo_dialog = OpenPhotoDialog()
        photo_filenames = open_photo_dialog.getOpenFileNames(self.main_window, 'Открыть фото', '',
                                                             "Изображения (*.jpg *.jpeg *.png)")
        if len(photo_filenames[0]) > 0:
            #print(photo_filenames[0])
            self.filenames = photo_filenames[0]
            self.photos_count = len(self.filenames)
            self.photos_selected = True
            self.adjustProcess()
        else:
            MyErrorsWarnings().open_photo_empty_warnings(self.main_window)

    def adjustProcess(self):
        adjusting_dialog = AdjustmentDialog(self.filenames)
        if adjusting_dialog.exec():
            self.conf = adjusting_dialog.getConf()
            self.iou = adjusting_dialog.getIoU()
            self.max_det = adjusting_dialog.getMaxDet()
            self.max_rows = adjusting_dialog.getRows()
            self.max_columns = adjusting_dialog.getColumns()
            print(self.conf)
            types = adjusting_dialog.getTypes()
            print("adjust ok", types)
            for i, image in enumerate(types):
                self.modes[image].append(self.filenames[i])
            self.radius_fractions = [self.config.cells_color_radius_fraction for i in range(self.photos_count)]
            self.setupUI()
            self.getInitialImages()
            self.findCells()
            self.visualizeResults()



    def setupUI(self):

        print("ok postprocess setupui")
        frame_stacked_widget = find_child(self.frame, child_type=QtWidgets.QStackedWidget)
        if frame_stacked_widget is not None:
            frame_stacked_widget.setParent(None)
            print("closing")

        # header
        self.header.setLineWidth(0)
        self.header.setFrameStyle(QtWidgets.QFrame.Shape.Panel)
        self.header.setFixedHeight(100)


        # ссылка на стек-виджет хидера
        header_stacked_widget = find_child(self.header, child_type=QtWidgets.QStackedWidget)

        # блок-контейнер для виджетов управления этого раздела
        header_photo_processing_box = find_child(header_stacked_widget, child_name="header_photo_processing_box")
        header_stacked_widget.setCurrentWidget(header_photo_processing_box)



        header_photo_processing_files_combobox = find_child(header_photo_processing_box, QtWidgets.QComboBox)
        header_photo_processing_files_combobox.clear()
        header_photo_processing_files_combobox.addItems(self.filenames)
        header_photo_processing_files_combobox.currentIndexChanged.connect(self.photosComboboxChanged)

        self.stages_button_group = find_child(header_photo_processing_box, QtWidgets.QButtonGroup)
        self.stages_button_group.idClicked.connect(self.onStageButtonClicked)

        header_save_button = find_child(header_photo_processing_box, child_name="header_photo_processing_save_button")
        header_save_button.clicked.connect(self.onSaveButtonClicked)
        print("header ok")

        # right sidebar
        self.right_sidebar.setFixedWidth(200)
        self.right_sidebar_stacked_widget = find_child(self.right_sidebar, child_type=QtWidgets.QStackedWidget)
        #self.right_sidebar_stacked_widget.setVisible(True)
        print("right_sidebar_stacked_widget found")
        right_sidebar_photo_processing_stacked_box = find_child(self.right_sidebar_stacked_widget, QtWidgets.QStackedWidget)
        print("right_sidebar_photo_processing_stacked_box found")
        self.right_sidebar_stacked_widget.setCurrentWidget(right_sidebar_photo_processing_stacked_box)
        print("right_sidebar_stacked_widget set")
        #right_sidebar_photoprocessing_info_frame = find_child(child_name="right_sidebar_photoprocessing_info_frame")
        right_sidebar_photoprocessing_contours_frame = find_child(right_sidebar_photo_processing_stacked_box, child_name="right_sidebar_photoprocessing_contours_frame")
        print("right_sidebar_photoprocessing_contours_frame found")
        #right_sidebar_photoprocessing_contours_layout = QtWidgets.QGridLayout()
        #right_sidebar_photoprocessing_contours_edit_button_icon = QtGui.QIcon("icons/icons8-select-cursor-24.png")
        #right_sidebar_photoprocessing_contours_edit_button = QtWidgets.QPushButton(right_sidebar_photoprocessing_contours_edit_button_icon, "Редактировать")
        right_sidebar_photoprocessing_contours_edit_button = find_child(right_sidebar_photoprocessing_contours_frame, child_name="right_sidebar_photoprocessing_contours_edit_button")
        print("right_sidebar_photoprocessing_contours_edit_button found")
        right_sidebar_photoprocessing_contours_edit_button.clicked.connect(self.onContoursEditButtonClicked)
        print("right_sidebar_photoprocessing_contours_edit_button connected")
        #right_sidebar_photoprocessing_contours_apply_button = QtWidgets.QPushButton("Применить")
        #right_sidebar_photoprocessing_contours_radius_fraction_input = find_child(right_sidebar_photoprocessing_contours_frame, child_type=QtWidgets.QDoubleSpinBox)
        # if right_sidebar_photoprocessing_contours_radius_fraction_input is not None:
        #     right_sidebar_photoprocessing_contours_radius_fraction_input.setParent(None)
        #     print("double spinbox closing")
        print("radius fractions: ", self.radius_fractions)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input = find_child(right_sidebar_photoprocessing_contours_frame, child_type=QtWidgets.QDoubleSpinBox)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setMinimum(0.1)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setMaximum(0.9)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setSingleStep(0.1)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setValue(self.config.cells_color_radius_fraction)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setReadOnly(True)
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.valueChanged.connect(self.onRadiusFractionSpinboxChanged)
        print(self.right_sidebar_photoprocessing_contours_radius_fraction_input, self.right_sidebar_photoprocessing_contours_radius_fraction_input.isReadOnly(), self.right_sidebar_photoprocessing_contours_radius_fraction_input.value())
        # right_sidebar_photoprocessing_contours_layout.addWidget(right_sidebar_photoprocessing_contours_edit_button, 1, 1)
        # right_sidebar_photoprocessing_contours_layout.addWidget(right_sidebar_photoprocessing_contours_apply_button, 1, 2)
        # right_sidebar_photoprocessing_contours_layout.addWidget(self.right_sidebar_photoprocessing_contours_radius_fraction_input, 2, 1)
        # right_sidebar_photoprocessing_contours_frame.setLayout(right_sidebar_photoprocessing_contours_layout)
        # right_sidebar_photoprocessing_colors_frame = QtWidgets.QFrame(objectName="right_sidebar_photoprocessing_contours_frame")
        # right_sidebar_photo_processing_stacked_box.addWidget(right_sidebar_photoprocessing_info_frame)
        # right_sidebar_photo_processing_stacked_box.addWidget(right_sidebar_photoprocessing_contours_frame)
        # right_sidebar_photo_processing_stacked_box.addWidget(right_sidebar_photoprocessing_colors_frame)
        # right_sidebar_photo_processing_stacked_box.setCurrentIndex(2)


        # frame
        self.frame_stacked_widget = QtWidgets.QStackedWidget(objectName="frame_stacked_widget")
        #self.frame_stacked_widget.setSizePolicy(QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Minimum)

        print("stacked widget init", self.frame_stacked_widget)
        #print(self.frame_stacked_widget.sizePolicy().horizontalPolicy())

        for i in range(self.photos_count):
            stage_stacked_widget = QtWidgets.QStackedWidget()
            #stage_stacked_widget.setSizePolicy(QtWidgets.QSizePolicy.Policy.Minimum,
            #                                        QtWidgets.QSizePolicy.Policy.Minimum)


            rectangle_graphics_view = QtWidgets.QGraphicsView()
            rectangle_graphics_scene = QtWidgets.QGraphicsScene()
            rectangle_graphics_view.setScene(rectangle_graphics_scene)

            contours_graphics_view = QtWidgets.QGraphicsView()
            contours_graphics_scene = QtWidgets.QGraphicsScene()
            contours_graphics_view.setScene(contours_graphics_scene)

            colors_graphics_view = QtWidgets.QGraphicsView()
            colors_graphics_scene = QtWidgets.QGraphicsScene()
            colors_graphics_view.setScene(colors_graphics_scene)

            stage_stacked_widget.addWidget(rectangle_graphics_view)
            stage_stacked_widget.addWidget(contours_graphics_view)
            stage_stacked_widget.addWidget(colors_graphics_view)
            stage_stacked_widget.setCurrentIndex(2)

            self.frame_stacked_widget.addWidget(stage_stacked_widget)

        frame_layout = find_child(self.frame, child_type=QtWidgets.QHBoxLayout)
        print("find child ok")
        frame_layout.addWidget(self.frame_stacked_widget)


        print(header_photo_processing_box.children())


    def updateContours(self, image_index):
        pass

    def updateColors(self, image_index):
        pass

    def photosComboboxChanged(self, ind):
        self.current_image_index = ind
        self.frame_stacked_widget.setCurrentIndex(ind)
        for button in self.stages_button_group.buttons():
            if self.stages_button_group.id(button) == self.frame_stacked_widget.widget(self.current_image_index).currentIndex():
                button.setChecked(True)
                break

        # right sidebar
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setValue(self.radius_fractions[ind])

    def onStageButtonClicked(self, ind):

        self.frame_stacked_widget.widget(self.current_image_index).setCurrentIndex(ind)
        self.right_sidebar_stacked_widget.currentWidget().setCurrentIndex(ind)

    def onContoursEditButtonClicked(self):
        self.right_sidebar_photoprocessing_contours_radius_fraction_input.setReadOnly(False)
        contours_graphics_view = self.frame_stacked_widget.widget(self.current_image_index).widget(1)
        current_items = contours_graphics_view.items()
        for item in current_items:
            if type(item) == QtWidgets.QGraphicsPixmapItem:
                current_pixmap_item = item
                break
        print(current_pixmap_item)
        print("onContoursEditButtonClicked")


    def onRadiusFractionSpinboxChanged(self, value):
        self.radius_fractions[self.current_image_index] = value



    def onSaveButtonClicked(self):
        self.write_simple_table()

    def drawContoursPixmap(self):
        pass

    def write_simple_table(self):
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Сохранить таблицу цветов", "",
                                                            "Excel Files (*.xlsx)", )
        print(filename)
        if filename:
            pass
        else:
            return
        wb = Workbook()
        abc = "ABCDEFGHIJKLMNQOPRST"
        for i, image in enumerate(self.filenames):
            ws = wb.create_sheet(self.filenames[i].split("/")[-1])
            c = ws.cell(row=1, column=2)
            c.value = "R"
            c = ws.cell(row=1, column=3)
            c.value = "G"
            c = ws.cell(row=1, column=4)
            c.value = "B"
            rows_count = self.cellspostprocess_data[image].rows_count
            columns_count = self.cellspostprocess_data[image].columns_count
            table_of_centers = self.cellspostprocess_data[image].table_of_centers
            table_of_blue_int = self.cellspostprocess_data[image].table_of_blue_int
            table_of_green_int = self.cellspostprocess_data[image].table_of_green_int
            table_of_red_int = self.cellspostprocess_data[image].table_of_red_int
            for j, row in enumerate(self.cellspostprocess_data[image].table_of_centers):
                for k, column in enumerate(row):
                    c = ws.cell(row=2+j*columns_count+k, column=1)
                    c.value = abc[j]+str(k+1)
                    c = ws.cell(row=2+j*columns_count+k, column=2)
                    c.value = table_of_red_int[table_of_centers[j][k]]
                    c = ws.cell(row=2 + j * columns_count + k, column=3)
                    c.value = table_of_green_int[table_of_centers[j][k]]
                    c = ws.cell(row=2 + j * columns_count + k, column=4)
                    c.value = table_of_blue_int[table_of_centers[j][k]]
        wb.save(filename)




    def write_old_table(self):

        filename, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Сохранить таблицу цветов", "", "Excel Files (*.xlsx)", )
        print(filename)
        if filename:
            pass
        else:
            return
        wb = Workbook()

        print(self.colors)
        abc = "ABCDEFGHIJKLMN"
        for k, sheet in enumerate(self.colors):
            ws = wb.create_sheet(self.filenames[k].split("/")[-1])
            print("ok")
            for i, row in enumerate(self.colors[k]):

                c = ws.cell(row=1, column=i*5+2)
                c.value = "R"

                c = ws.cell(row=1, column=i * 5 + 3)
                c.value = "G"
                c = ws.cell(row=1, column=i * 5 + 4)
                c.value = "B"

                for j, rgb in enumerate(row):
                    r, g, b = rgb
                    c = ws.cell(row=j+2, column=i*5+1)
                    c.value = abc[i]+str(j+1)
                    c = ws.cell(row=j+2, column=i*5+2)
                    c.value = r
                    c = ws.cell(row=j + 2, column=i * 5 + 3)
                    c.value = g
                    c = ws.cell(row=j + 2, column=i * 5 + 4)
                    c.value = b


        wb.save(filename)

    def getInitialImages(self):
        for image in self.filenames:
            file = open(image, "rb")
            data = file.read()
            initial_image = cv2.imdecode(np.frombuffer(data, np.uint8), flags=cv2.IMREAD_COLOR)
            #initial_image = cv2.cvtColor(initial_image, cv2.COLOR_BGR2RGB)
            self.initial_images[image] = initial_image

    def findCells(self):
        for mode in self.modes.keys():
            if len(self.modes[mode]) > 0:
                cv_images = [self.initial_images[n] for n in self.modes[mode]]
                names = self.modes[mode]
                cf = CellsFinder(cv_images, mode, conf=self.conf, iou=self.iou, max_det=self.max_det)
                print(f"mode {mode} is done!")
                for i, name in enumerate(names):
                    self.results[name] = cf.results[i]


    def _findCells(self):
        #modes = {'rgb camag': 0, 'rgb camera': 0, 'ir': 0, '366': 0, '254': 0}
        for i, image in enumerate(self.filenames):
            mode = os.path.splitext(image)[0].split(sep=" ")[-1]
            if mode in self.modes.keys():
                self.modes[mode].append(image)
                self.images_modes[image] = mode
            else:
                self.images_modes[image] = mode
                self.wrong_modes.append(mode)
                self.wrong_images.append(image)
        # сделать мультипроцессорное распознавание для разных типов
        for mode in self.modes.keys():
            if len(self.modes[mode]) > 0:
                cv_images = [self.initial_images[n] for n in self.modes[mode]]
                names = self.modes[mode]
                cf = CellsFinder(cv_images, mode)
                print(f"mode {mode} is done!")
                for i, name in enumerate(names):
                    self.results[name] = cf.results[i]
        if len(self.wrong_images) > 0:
            cv_images = [self.initial_images[n] for n in self.wrong_images]
            cf = CellsFinder(cv_images, 'universal')
            print("mode universal is done")
            for i, n in enumerate(self.wrong_images):
                self.results[n] = cf.results[i]
        print("ok cf")







    def visualizeResults(self):
        print("visualize start")
        for i, name_image in enumerate(self.initial_images.items()):
            cp = CellsPostprocessor(self.results[name_image[0]], name_image[1], self.max_rows, self.max_columns)
            self.cellspostprocess_data[name_image[0]] = cp
            print("ok cp")


            initial_graphics_view = self.frame_stacked_widget.widget(i).widget(0)
            contours_graphics_view = self.frame_stacked_widget.widget(i).widget(1)
            colors_graphics_view = self.frame_stacked_widget.widget(i).widget(2)



            scene_width = self.frame.width()-30
            scene_height = self.frame.height()-30
            scene_aspect_ratio = scene_width / scene_height
            if cp.success:
                image_aspect_ratio = cp.rotated_image.shape[1]/cp.rotated_image.shape[0] #для повернутого изображения
                print(image_aspect_ratio, scene_aspect_ratio)
            else:
                image_aspect_ratio = cp.image.shape[1]/cp.image.shape[0]
            if image_aspect_ratio < scene_aspect_ratio:
                resized_width = round(scene_height * image_aspect_ratio)
                resized_height = scene_height

            else:
                resized_width = scene_width
                resized_height = round(scene_width / image_aspect_ratio)
            #pix_initial_image = from_cv2_to_pixmap(scale_cv2(cp.rotated_image, resized_width, resized_height))
            if cp.success:
                pix_initial_image = from_cv2_to_pixmap(cp.rotated_image)
            else:
                pix_initial_image = from_cv2_to_pixmap(cp.image)
            #initial_graphics_view.setSceneRect(0, 0, 1200, 600)
            #initial_graphics_view.fitInView(initial_graphics_view.scene().sceneRect())
            #initial_graphics_view.setSceneRect(0, 0, 700, 600)
            initial_graphics_view.scene().addPixmap(pix_initial_image.scaled(resized_width, resized_height, transformMode=QtCore.Qt.TransformationMode.SmoothTransformation))
            #initial_graphics_view.scene().addPixmap(pix_initial_image)
            print("ok init resize")
            print("frame", self.frame.width(), self.frame.height())
            print("frame stacked widget", self.frame_stacked_widget.size(), self.frame_stacked_widget.sizePolicy())
            print("graphics stacked widget", self.frame_stacked_widget.widget(i).width(), self.frame_stacked_widget.widget(i).height())
            print("graphics view", initial_graphics_view.sceneRect(), initial_graphics_view.width(), initial_graphics_view.height())
            print("scene", initial_graphics_view.scene().width(), initial_graphics_view.scene().height())
            print("pix", pix_initial_image.width(), pix_initial_image.height())
            if cp.success:
                image_aspect_ratio = cp.contours_image.shape[1]/cp.contours_image.shape[0] #для повернутого изображения
                print(image_aspect_ratio, scene_aspect_ratio)
            else:
                image_aspect_ratio = cp.image.shape[1]/cp.image.shape[0]
            if image_aspect_ratio < scene_aspect_ratio:
                resized_width = round(scene_height * image_aspect_ratio)
                resized_height = scene_height
            else:
                resized_width = scene_width
                resized_height = round(scene_width / image_aspect_ratio)
            if cp.success:
                pix_contours_image = from_cv2_to_pixmap(cp.contours_image, )
            else:
                pix_contours_image = from_cv2_to_pixmap(cp.image)
            contours_pixmap_item = contours_graphics_view.scene().addPixmap(pix_contours_image.scaled(resized_width, resized_height, transformMode=QtCore.Qt.TransformationMode.SmoothTransformation))
            #contours_graphics_view.scene().addPixmap(pix_contours_image)
            print("ok cont resize")
            if not cp.success:
                continue
            rows_count = cp.rows_count
            columns_count = cp.columns_count
            colors_aspect_ratio = columns_count / rows_count


            rgb_pretty_image = cv2.cvtColor(cp.pretty_image, cv2.COLOR_BGR2RGB)
            print("ok bgr to rgb")
            pix_colors_image = from_cv2_to_pixmap(cp.pretty_image)
            print("ok colors cv to pix")
            image_aspect_ratio = pix_colors_image.width() / pix_colors_image.height()  # для повернутого изображения
            if image_aspect_ratio < scene_aspect_ratio:
                resized_width = round(scene_height * image_aspect_ratio)
                resized_height = scene_height

            else:
                resized_width = scene_width
                resized_height = round(scene_width / image_aspect_ratio)

            colors_graphics_view.scene().addPixmap(pix_colors_image.scaled(resized_width, resized_height, transformMode=QtCore.Qt.TransformationMode.SmoothTransformation))
            #colors_graphics_view.scene().addPixmap(pix_colors_image)
            print("ok colors resize")








    def preprocessImages(self):
        mode = " ".join(os.path.splitext(self.filenames[0])[0].split(sep=" ")[1:])
        cf = CellsFinder(self.filenames, mode)
        for i in range(self.photos_count):
            rectangle_graphics_view = self.frame_stacked_widget.widget(i).widget(0)
            contours_graphics_view = self.frame_stacked_widget.widget(i).widget(1)
            colors_graphics_view = self.frame_stacked_widget.widget(i).widget(2)


            # открыть изображение с помощью OpenCV
            # open image with OpenCV
            start = time.time()

            file = open(self.filenames[i], "rb")
            data = file.read()
            initial_image = cv2.imdecode(np.frombuffer(data, np.uint8), flags=cv2.IMREAD_COLOR)
            initial_image = cv2.cvtColor(initial_image, cv2.COLOR_BGR2RGB)
            self.initial_images.append(initial_image)
            pix = from_cv2_to_pixmap(initial_image)
            print(pix.width(), pix.height())
            scene_width = rectangle_graphics_view.width()
            scene_height = rectangle_graphics_view.height()
            print(scene_width, scene_height)
            scale_w = int(scene_width / pix.width())
            scale_h = int(scene_height / pix.height())
            pix2 = pix.scaled(scene_width, scene_height)
            if scale_h < scale_w:
                pix1 = pix.scaledToWidth(scene_width)
            else:
                pix1 = pix.scaledToHeight(scene_height)
            print(pix1.width(), pix1.height())
            rectangle_graphics_view.scene().addPixmap(pix1)


            print("Initial image done", time.time()-start)

            # попытаься найти на фото непосредственно планшетку
            # try to find only tablet
            #
            find_main_rectangle_class = Finder_of_main_rectangle(initial_image)
            print("Main rect done", time.time() - start)
            if find_main_rectangle_class is not None:
                # установить полученное изображение планшетки
                # set tablet image as image for further work
                image = find_main_rectangle_class.main_rectangle_image

            else:
                # иначе искать контуры на исходном фото
                # use initial image
                image = initial_image


            print("ok")
            find_contours_class = Finder_of_circles_contours_cv2(image)
            contours_image = np.copy(image)
            cv2.drawContours(contours_image, find_contours_class.contours, -1, (0,255,255), 3)
            contours_pix = from_cv2_to_pixmap(contours_image)
            contours_graphics_view.scene().addPixmap(contours_pix)


            print("Contours done", time.time() - start)
            try:
                checker_class = Checker_of_circles_contours(image, find_contours_class.contours)
                print(checker_class.table_of_centers)
                print("checker")
            except:
                print("continue")
                continue

            table_of_centers = checker_class.table_of_centers
            print("Table done", time.time() - start)
            colors_class = ColorExtracter(image, table_of_centers, checker_class.radius)
            print("Colors done", time.time() - start)
            print(colors_class.colors)

            self.colors.append(colors_class.colors)
            colors_image = np.copy(image)
            colors_pix = from_cv2_to_pixmap(colors_image)
            #colors_graphics_view.scene().addPixmap(colors_pix)
            for i, row in enumerate(table_of_centers):
                for j, circle in enumerate(row):
                    #print(*colors_class.colors[i][j])
                    brush = QtGui.QBrush(QtGui.QColor(*colors_class.colors[i][j]))
                    #print("brush ok")
                    ellipse = QtWidgets.QGraphicsEllipseItem(*circle[1], 2*checker_class.radius, 2*checker_class.radius)
                    #print("ellipse ok")
                    ellipse.setBrush(brush)
                    #print("set ok")
                    colors_graphics_view.scene().addItem(ellipse)
                    #print("add ok")
    def close(self):

        header_stacked_widget = find_child(self.header, child_type=QtWidgets.QStackedWidget)




        # open_photo_dialog = find_child(self.main_window, child_type=OpenPhotoDialog)
        # if open_photo_dialog is not None:
        #     open_photo_dialog.cl

class MyErrorsWarnings:
    def __init__(self):
        pass

    def open_photo_empty_warnings(self, parent):
        dlg = QtWidgets.QMessageBox(parent)
        dlg.setWindowTitle("Сообщение")
        dlg.setText("Фото не выбрано")
        button = dlg.exec()





def main():



    app = QtWidgets.QApplication(sys.argv)

    w = MainWindow()
    app.setStyleSheet(style)
    app.exec()


if __name__ == "__main__":
    main()
